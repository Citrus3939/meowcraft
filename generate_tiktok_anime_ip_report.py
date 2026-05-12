from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_DATE = "2026-05-12"
OUT_DIR = Path("reports")
OUT_DIR.mkdir(exist_ok=True)

TIKTOK_HASHTAGS_BASE = "https://tiktokhashtags.com/hashtag/{tag}/"
CONNECT_TIKTOK_DEMO = "https://connectmgt.com/agency/blog/tiktok-demographics-2025-whos-using-the-platform-today"
DATALOOK_TIKTOK_DEMO = "https://datalook.io/statistics/tiktok-users"
ANN_CRUNCHYROLL_NRG = "https://www.animenewsnetwork.com/interest/2025-05-23/crunchyroll-research-over-half-of-gen-z-globally-are-anime-fans/.224732"
FANDOM_ANIME_2023 = "https://about.fandom.com/insights/inside-anime-2023"
DENTSU_REPORT_SEARCH = "Dentsu Anime - A Growing Opportunity for Brands (2025 Global Research Report), public search snippet"


records = [
    {
        "ip": "Naruto / 火影忍者",
        "tag": "naruto",
        "posts_wan": 3130.0,
        "views_yi": 3305.0,
        "profile": "老牌热血王道 IP；TikTok 上以名场面剪辑、忍术梗、角色战力争议、cosplay 和怀旧二创为主。",
        "gender": "男55-65% / 女35-45%（估算）",
        "age": "13-24 为传播核心；25-34 怀旧复看与购买力更强",
        "spending": "高：手办、服饰、游戏、卡牌/周边、漫展 cosplay 均有成熟消费场景",
        "opportunity": "角色剪辑、忍者动作挑战、服饰联名、游戏/手办转化",
        "method": "公开原始热度=主标签 #naruto；画像=题材+TikTok/动漫总体报告估算",
    },
    {
        "ip": "One Piece / 海贼王",
        "tag": "onepiece",
        "posts_wan": 1590.0,
        "views_yi": 2549.0,
        "profile": "全球长线连载与真人剧加持；内容集中在角色高光、剧情解读、恶魔果实设定、船员情感和 cosplay。",
        "gender": "男55-62% / 女38-45%（估算）",
        "age": "13-34 覆盖最宽；18-34 为主消费层",
        "spending": "极高：漫画、流媒体、手办、服饰、游戏、联名商品和线下活动齐全",
        "opportunity": "剧情解读、角色混剪、真人剧联动、服饰/潮玩转化",
        "method": "公开原始热度=主标签 #onepiece；画像=题材+长线 IP 受众估算",
    },
    {
        "ip": "Demon Slayer / 鬼灭之刃",
        "tag": "demonslayer",
        "posts_wan": 1200.0,
        "views_yi": 1197.0,
        "profile": "视觉强、角色辨识度高；TikTok 重点内容是炭治郎/祢豆子/柱角色剪辑、变装、cosplay 和名场面 BGM。",
        "gender": "男45-55% / 女45-55%（估算）",
        "age": "13-24 传播最强；亲子/大众层也有外溢",
        "spending": "高：角色周边、服饰、刀具道具、盲盒、电影/流媒体消费强",
        "opportunity": "短变装、角色道具、BGM 卡点、家庭向与二次元周边",
        "method": "公开原始热度=主标签 #demonslayer；未叠加 #kimetsunoyaiba 以避免重复计算",
    },
    {
        "ip": "My Hero Academia / 我的英雄学院",
        "tag": "mha",
        "posts_wan": 1390.0,
        "views_yi": 916.0,
        "profile": "校园英雄群像适合角色站队、CP/同人、能力设定讨论；TikTok 上剪辑和二创活跃。",
        "gender": "男45-55% / 女45-55%（估算）",
        "age": "13-24 为核心；18-24 互动与同人创作更活跃",
        "spending": "中高：角色立牌、服装、手办、漫画和漫展 cosplay 消费明显",
        "opportunity": "角色阵营话题、能力测试、校园英雄变装、同人周边",
        "method": "公开原始热度=别名主标签 #mha；#myheroacademia 另有约686亿浏览",
    },
    {
        "ip": "Dragon Ball / 龙珠",
        "tag": "dragonball",
        "posts_wan": 660.0,
        "views_yi": 892.0,
        "profile": "男性向与怀旧属性强；内容以战斗名场面、悟空/贝吉塔梗、健身热血剪辑和游戏片段为主。",
        "gender": "男65-75% / 女25-35%（估算）",
        "age": "18-34 最强；13-17 通过游戏和短剪辑进入",
        "spending": "高：手办、模型、游戏、服饰和收藏品消费基础深",
        "opportunity": "热血训练挑战、战力榜、游戏联动、复古周边",
        "method": "公开原始热度=主标签 #dragonball；未叠加 #dbz/#dragonballz",
    },
    {
        "ip": "Attack on Titan / 进击的巨人",
        "tag": "aot",
        "posts_wan": 740.0,
        "views_yi": 841.0,
        "profile": "剧情反转、史诗感和角色悲剧性适合高完播剪辑；受众偏成熟，讨论密度高。",
        "gender": "男55-65% / 女35-45%（估算）",
        "age": "18-34 为核心；13-17 通过名场面剪辑进入",
        "spending": "高：收藏手办、服饰、海报、漫画典藏和高客单模型有空间",
        "opportunity": "名场面剪辑、剧情解析、角色立场辩论、暗黑服饰联名",
        "method": "公开原始热度=别名 #aot；#attackontitan 另有约748亿浏览",
    },
    {
        "ip": "Jujutsu Kaisen / 咒术回战",
        "tag": "jujutsukaisen",
        "posts_wan": 590.0,
        "views_yi": 827.0,
        "profile": "五条悟、宿傩等角色自带流量；TikTok 适配角色颜值剪辑、战斗卡点、cosplay 和 meme。",
        "gender": "男50-58% / 女42-50%（估算）",
        "age": "13-24 最强；18-24 购买力和二创活跃度高",
        "spending": "高：角色周边、服饰、手办、漫画和 cosplay 道具转化好",
        "opportunity": "五条悟/宿傩角色内容、战斗卡点、潮流服饰联名",
        "method": "公开原始热度=主标签 #jujutsukaisen；未叠加 #jjk",
    },
    {
        "ip": "Haikyuu!! / 排球少年",
        "tag": "haikyuu",
        "posts_wan": 760.0,
        "views_yi": 778.0,
        "profile": "运动热血与角色关系驱动；TikTok 上以角色混剪、队伍阵营、情绪向剪辑和同人内容为主。",
        "gender": "男35-45% / 女55-65%（估算）",
        "age": "13-24 为传播核心；18-24 同人和周边购买更强",
        "spending": "中高：角色周边、立牌、服饰、海报、运动相关联名适配度高",
        "opportunity": "角色关系剪辑、球队站队、运动服饰和校园场景合作",
        "method": "公开原始热度=主标签 #haikyuu；画像按运动番+同人活跃特征估算",
    },
    {
        "ip": "Pokémon / 宝可梦",
        "tag": "pokemon",
        "posts_wan": 950.0,
        "views_yi": 722.0,
        "profile": "全年龄、游戏和收藏属性极强；TikTok 内容覆盖卡牌开箱、游戏、可爱宠物梗、怀旧和动画角色。",
        "gender": "男50-55% / 女45-50%（估算）",
        "age": "13-34 覆盖宽；18-34 收藏和游戏付费更强",
        "spending": "极高：卡牌、游戏、毛绒、联名、收藏品和线下活动消费链完整",
        "opportunity": "卡牌开箱、萌宠化二创、游戏攻略、潮玩/零售联名",
        "method": "公开原始热度=主标签 #pokemon；作为动画/游戏复合 IP 处理",
    },
    {
        "ip": "Tokyo Revengers / 东京复仇者",
        "tag": "tokyorevengers",
        "posts_wan": 580.0,
        "views_yi": 716.0,
        "profile": "不良少年、时间穿越和高颜值角色适合短视频二创；女性向角色消费与男性向热血叙事并存。",
        "gender": "男40-50% / 女50-60%（估算）",
        "age": "13-24 为核心；18-24 cosplay/服饰消费明显",
        "spending": "中高：服饰、角色周边、cosplay、海报和同人商品较强",
        "opportunity": "角色发型/服装模仿、帮派阵营话题、情绪向剪辑",
        "method": "公开原始热度=主标签 #tokyorevengers；未叠加 #tokyorevengersedit",
    },
    {
        "ip": "Boruto / 博人传",
        "tag": "boruto",
        "posts_wan": 260.0,
        "views_yi": 469.0,
        "profile": "依托 Naruto 世界观和新世代角色；内容以父子传承、战斗片段、角色争议和续作讨论为主。",
        "gender": "男60-70% / 女30-40%（估算）",
        "age": "13-24 新观众；25-34 Naruto 老粉参与讨论",
        "spending": "中：流媒体、漫画、游戏联动和轻周边为主，收藏强度低于 Naruto 本体",
        "opportunity": "新旧角色对比、战力争议、父子传承梗、游戏联动",
        "method": "公开原始热度=主标签 #boruto；未并入 Naruto 以保留独立标题热度",
    },
    {
        "ip": "JoJo's Bizarre Adventure / JOJO 的奇妙冒险",
        "tag": "jojo",
        "posts_wan": 330.0,
        "views_yi": 382.0,
        "profile": "强 meme、姿势、音乐和时尚美学；TikTok 上非常适合模仿、卡点和反差剪辑。",
        "gender": "男60-70% / 女30-40%（估算）",
        "age": "18-34 为核心；13-17 通过 meme 入坑",
        "spending": "中高：手办、服饰、海报、音乐梗周边和收藏品有稳定需求",
        "opportunity": "JOJO 立、BGM 卡点、时尚联名、meme 二创",
        "method": "公开原始热度=主标签 #jojo；#jojosbizarreadventure 另有约297亿浏览",
    },
    {
        "ip": "Bleach / 死神",
        "tag": "bleach",
        "posts_wan": 240.0,
        "views_yi": 344.0,
        "profile": "千年血战篇带动回流；TikTok 内容以卍解名场面、角色颜值、战力榜和怀旧剪辑为主。",
        "gender": "男60-70% / 女30-40%（估算）",
        "age": "18-34 为核心；25-34 老粉购买力更强",
        "spending": "高：手办、服饰、武器道具、漫画典藏和游戏联动空间大",
        "opportunity": "卍解卡点、角色战力榜、复古回忆杀、武器道具周边",
        "method": "公开原始热度=主标签 #bleach；画像按热血长线 IP 估算",
    },
    {
        "ip": "Hunter x Hunter / 全职猎人",
        "tag": "hxh",
        "posts_wan": 320.0,
        "views_yi": 312.0,
        "profile": "角色关系和念能力系统讨论强；TikTok 内容包括奇犽/小杰剪辑、战力分析和情绪向二创。",
        "gender": "男55-65% / 女35-45%（估算）",
        "age": "13-34 均有；18-34 讨论和收藏更强",
        "spending": "中高：角色手办、服饰、漫画、海报和复古周边有稳定需求",
        "opportunity": "念能力测试、角色羁绊剪辑、战力讨论、复古动漫周边",
        "method": "公开原始热度=别名 #hxh；#hunterxhunter 另有约270亿浏览",
    },
    {
        "ip": "Chainsaw Man / 电锯人",
        "tag": "chainsawman",
        "posts_wan": 170.0,
        "views_yi": 283.0,
        "profile": "暗黑、血腥、反英雄和角色性格强烈；TikTok 上以玛奇玛/帕瓦/电次剪辑、cosplay 和潮流梗为主。",
        "gender": "男55-65% / 女35-45%（估算）",
        "age": "18-24 最强；内容题材偏成熟，不建议按低龄核心投放",
        "spending": "高：潮流服饰、手办、海报、成人向收藏和 cosplay 消费强",
        "opportunity": "暗黑潮流联名、角色变装、BGM 卡点、收藏级周边",
        "method": "公开原始热度=主标签 #chainsawman；画像按暗黑青年向题材估算",
    },
    {
        "ip": "Spy x Family / 间谍过家家",
        "tag": "spyxfamily",
        "posts_wan": 88.12,
        "views_yi": 212.0,
        "profile": "阿尼亚可爱梗和家庭喜剧属性极强；TikTok 上适合表情包、亲子内容、萌系变装和轻剧情剪辑。",
        "gender": "男35-45% / 女55-65%（估算）",
        "age": "13-34 覆盖宽；女性、亲子和轻二次元受众更易进入",
        "spending": "高：毛绒、盲盒、文具、服饰、亲子/礼品消费转化好",
        "opportunity": "阿尼亚表情包、亲子礼品、萌系周边、轻喜剧短视频",
        "method": "公开原始热度=主标签 #spyxfamily；画像按萌系家庭喜剧特征估算",
    },
    {
        "ip": "Blue Lock / 蓝色监狱",
        "tag": "bluelock",
        "posts_wan": 150.0,
        "views_yi": 203.0,
        "profile": "足球竞技+高颜值角色群像；TikTok 上以角色剪辑、CP/同人、运动热血和世界杯/足球话题联动为主。",
        "gender": "男40-50% / 女50-60%（估算）",
        "age": "13-24 为核心；18-24 周边与同人消费活跃",
        "spending": "中高：角色立牌、球衣/运动服饰、海报、漫画和同人周边适配",
        "opportunity": "球员阵营、运动挑战、球衣联名、角色关系剪辑",
        "method": "公开原始热度=主标签 #bluelock；画像按运动番+女性向角色消费估算",
    },
    {
        "ip": "Death Note / 死亡笔记",
        "tag": "deathnote",
        "posts_wan": 220.0,
        "views_yi": 194.0,
        "profile": "智斗、暗黑美学和 L/月的角色对立长期出圈；TikTok 上以台词、推理梗、哥特审美和怀旧剪辑为主。",
        "gender": "男50-60% / 女40-50%（估算）",
        "age": "18-34 为核心；13-17 通过梗和短剪辑入坑",
        "spending": "中：服饰、海报、笔记本、漫画典藏和暗黑风周边更适配",
        "opportunity": "智斗梗、暗黑美学、笔记本道具、经典台词二创",
        "method": "公开原始热度=主标签 #deathnote；画像按经典暗黑智斗 IP 估算",
    },
    {
        "ip": "Black Clover / 黑色五叶草",
        "tag": "blackclover",
        "posts_wan": 78.78,
        "views_yi": 179.0,
        "profile": "少年热血与魔法战斗设定清晰；TikTok 上以战斗剪辑、主角逆袭、角色能力和热血 BGM 为主。",
        "gender": "男60-70% / 女30-40%（估算）",
        "age": "13-24 为核心；18-24 动漫剪辑消费更强",
        "spending": "中：漫画、轻周边、手办和流媒体消费为主，高客单收藏相对弱",
        "opportunity": "逆袭叙事、魔法能力排行、热血剪辑、轻周边投放",
        "method": "公开原始热度=主标签 #blackclover；画像按少年热血题材估算",
    },
    {
        "ip": "One Punch Man / 一拳超人",
        "tag": "onepunchman",
        "posts_wan": 86.09,
        "views_yi": 172.0,
        "profile": "反套路英雄和强 meme 属性；TikTok 上适合埼玉反差、健身梗、战斗高光和喜剧剪辑。",
        "gender": "男65-75% / 女25-35%（估算）",
        "age": "18-34 为核心；13-24 通过 meme 和战斗剪辑传播",
        "spending": "中：服饰、手办、海报和梗周边适合，整体消费链低于顶级长线 IP",
        "opportunity": "健身挑战、反差喜剧、英雄战力梗、轻量联名周边",
        "method": "公开原始热度=主标签 #onepunchman；画像按男性向喜剧热血题材估算",
    },
]


headers = [
    "排名",
    "动漫IP",
    "TikTok主标签",
    "公开浏览量(亿)",
    "公开帖子量(万)",
    "热度依据",
    "用户画像",
    "男女比例",
    "核心年龄段",
    "消费力",
    "消费品类机会",
    "数据口径/备注",
]


records = sorted(records, key=lambda row: row["views_yi"], reverse=True)

wb = Workbook()
ws = wb.active
ws.title = "Top20_TikTok动漫IP"
ws.append(headers)

for rank, row in enumerate(records, 1):
    ws.append(
        [
            rank,
            row["ip"],
            f"#{row['tag']}",
            row["views_yi"],
            row["posts_wan"],
            f"TikTokHashtags公开页：#{row['tag']} overall views/posts",
            row["profile"],
            row["gender"],
            row["age"],
            row["spending"],
            row["opportunity"],
            row["method"],
        ]
    )

source_ws = wb.create_sheet("数据来源与口径")
source_ws.append(["项目", "说明"])
source_rows = [
    ("采集日期", REPORT_DATE),
    ("生成时间UTC", datetime.now(UTC).isoformat(timespec="seconds")),
    ("排名口径", "按 TikTokHashtags 可公开访问的单一主标签 overall views 排序；不叠加别名标签，避免重复计算同一批视频。"),
    ("TikTok热度来源", "TikTokHashtags hashtag pages；页面示例：" + TIKTOK_HASHTAGS_BASE.format(tag="naruto")),
    ("性别/年龄说明", "TikTok 官方公开页不提供按动漫 IP 拆分的男女比例和年龄段；本表为画像估算，不是官方后台数据。"),
    ("TikTok总体人口统计参考1", "Connect Management：全球约1.59B用户；18-24约35%；13-17约18%；全球女性54.8%、男性45.2%。" + " " + CONNECT_TIKTOK_DEMO),
    ("TikTok总体人口统计参考2", "DataLook：女性56.2%、男性43.8%；美国18-24占25%、25-34占35%。" + " " + DATALOOK_TIKTOK_DEMO),
    ("动漫受众年龄参考", "Crunchyroll/NRG via Anime News Network：13-28岁Gen Z中54%为动漫粉，Millennials 42%，Gen X 24%。" + " " + ANN_CRUNCHYROLL_NRG),
    ("动漫消费参考1", "Dentsu 2025公开报告搜索片段：约28%全球动漫观看者购买喜爱IP周边；超过40%通过TikTok/Instagram fan edits发现新动漫。"),
    ("动漫消费参考2", "Fandom Inside Anime 2023：78%动漫粉每日/每周观看；60%+为游戏重度受众；动漫市场2027预计达440亿美元。" + " " + FANDOM_ANIME_2023),
    ("消费力定义", "极高=有游戏/卡牌/收藏/授权商品全链路；高=手办/服饰/cosplay/漫展转化强；中高=角色周边和服饰明显；中=轻周边和内容消费为主。"),
    ("缺失项", "未使用登录态、广告后台、TikTok私有接口或付费面板；没有单IP官方精确性别、年龄、收入或GMV数据。"),
]
for item, desc in source_rows:
    source_ws.append([item, desc])

raw_ws = wb.create_sheet("原始标签指标")
raw_ws.append(["动漫IP", "主标签", "帖子量(万)", "浏览量(亿)", "来源URL"])
for row in records:
    raw_ws.append(
        [
            row["ip"],
            f"#{row['tag']}",
            row["posts_wan"],
            row["views_yi"],
            TIKTOK_HASHTAGS_BASE.format(tag=row["tag"]),
        ]
    )

for sheet in wb.worksheets:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value[:80]) + 2), 65)
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

out_path = OUT_DIR / "tiktok_anime_ip_top20_audience.xlsx"
wb.save(out_path)
print(out_path.resolve())
