from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_DATE = "2026-05-12"
OUT_DIR = Path("reports")
OUT_DIR.mkdir(exist_ok=True)
DESKTOP = Path.home() / "Desktop"
DESKTOP.mkdir(exist_ok=True)

HYPEAUDITOR = "https://hypeauditor.com/instagram-category-cosplay/"
FEEDSPOT = "https://influencers.feedspot.com/cosplay_instagram_influencers/"
INFLDB = "https://infldb.com/top-influencers-by-category/cosplay"
HIREINFLUENCE = "https://hireinfluence.com/blog/top-cosplay-influencers/"
COLLABSTR = "https://collabstr.com/top-influencers/cosplay"
SEARCH = "WebSearch result snippets captured on 2026-05-12"


MISSING = "[缺失；尝试3次]"
MISSING_WORK = "[缺失；尝试3次：公开来源未列最高播放/点赞角色]"


def rec(
    name,
    handle,
    platform,
    followers,
    style,
    monetization,
    reason,
    source,
    er=None,
    growth=None,
    works=None,
    notes="",
):
    return {
        "name": name,
        "handle": handle,
        "platform": platform,
        "followers": followers,
        "er": er,
        "growth": growth,
        "style": style,
        "monetization": monetization,
        "works": works or MISSING_WORK,
        "reason": reason,
        "source": source,
        "notes": notes,
    }


records = [
    rec("KNITE", "@knitecoser", "双平台", 100.0, "欧美游戏/影视；道具/机械装甲", "广告合作；自营店铺", "因为技术壁垒+IP红利+互动人设，游戏/动漫角色与高规格造型形成跨平台辨识度。", SEARCH, er=19.58, notes="Heepsy公开片段：1M followers, 19.58% engagement。"),
    rec("cosplayshorty00", "@cosplayshorty00", "TikTok", 50.02, "日系动漫；色气/性感风", "广告合作", "因为高频整活/短变装节奏+强力视觉标签，短视频更新密度与角色化包装更适合TikTok扩散。", SEARCH, er=12.61, notes="Urlebird公开片段：500.22K followers, 12.61% video engagement。"),
    rec("Komori Cosplay", "@komori198", "IG", 38.10, "日系动漫；色气/性感风", "广告合作；自营店铺", "因为强力原创人设/视觉美学标签+IP红利，稳定输出动漫向角色图集并导流服装合作。", HYPEAUDITOR, er=7.79),
    rec("PINKU cosplay", "@pinku.cosplay", "IG", 24.68, "日系动漫；色气/性感风", "广告合作", "因为强力原创人设/视觉美学标签+IP红利，粉色系动漫造型和高互动图文形成记忆点。", HYPEAUDITOR, er=10.03),
    rec("Fabibi World Cosplay", "@fabibiworldcosplay", "IG", 62.30, "日系动漫；漫展现实感", "广告合作", "因为IP红利+互动人设，cosplay与旅行/线下场景结合，内容具备生活化传播入口。", HYPEAUDITOR, er=4.21, works="[缺失；尝试3次：Feedspot/HypeAuditor未列最高互动角色]"),
    rec("Cosplay | Anime | Cosplaygirl", "@bellascosplay", "IG", 37.22, "日系动漫；色气/性感风", "广告合作", "因为IP红利+强力视觉标签，账号长期聚焦动漫cosplay垂类并获得高频互动。", HYPEAUDITOR, er=6.01),
    rec("Nabi", "@rhauura", "IG", 23.33, "日系动漫；洛丽塔/哥特/暗黑美学", "广告合作", "因为强力原创人设/视觉美学标签+IP红利，暗黑二次元审美让账号在同类中更易被识别。", HYPEAUDITOR, er=5.17),
    rec("Uwowo Cosplay Official", "@uwowo.cosplay", "IG", 20.39, "日系动漫；色气/性感风", "自营店铺；广告合作", "因为IP红利+强力视觉标签，服装品牌账号借热门动漫角色与成衣转化形成增长闭环。", HYPEAUDITOR, er=5.45),
    rec("Ribaibu Cosplay", "@ribaibu", "IG", 13.23, "日系动漫；漫展现实感", "广告合作", "因为IP红利+互动人设，coser个人形象和漫展/角色内容共同驱动粉丝黏性。", HYPEAUDITOR, er=5.34),
    rec("frailu", "@frailu", "IG", 12.20, "日系动漫；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，个人coser身份与稳定图文互动形成小而稳的圈层影响力。", SEARCH, er=1.538, notes="StarNgage公开片段：122K followers, 1.538% engagement。"),
    rec("mel", "@melondoki", "IG", 9.58, "日系动漫；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+高频短内容节奏，个人审美统一且互动率高。", HYPEAUDITOR, er=8.65),
    rec("Zali", "@zalariacos", "IG", 8.89, "日系动漫；欧美游戏/影视", "广告合作", "因为IP红利+互动人设，Fire Emblem等游戏向角色内容面向高黏性粉丝圈层。", SEARCH, er=3.8, notes="SocialVeins/StarNgage公开片段：88.9K followers, 3.8% engagement。"),
    rec("Julha | Cosplayer", "@juhlets", "IG", 7.75, "日系动漫；漫展现实感", "广告合作", "因为互动人设+IP红利，cosplayer身份清晰且通过角色/漫展内容维持高互动。", HYPEAUDITOR, er=8.48),
    rec("Cosplay Indonesia - Cosclip ID", "@cosclip.id", "IG", 7.53, "漫展现实感；日系动漫", "广告合作", "因为互动人设+IP红利，地区性cosplay社群账号聚合线下漫展与动漫角色内容。", HYPEAUDITOR, er=5.8),
    rec("simrell", "@katiesimrell", "IG", 7.18, "原创/OC；日系动漫", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，小体量账号依靠鲜明个人形象拿到高互动率。", HYPEAUDITOR, er=7.82),
    rec("Kseniya", "@kamiko_zero", "IG", 7.05, "日系动漫；洛丽塔/哥特/暗黑美学", "广告合作", "因为强力视觉标签+IP红利，俄语圈二次元造型与稳定写真内容带来高互动。", HYPEAUDITOR, er=7.21),
    rec("Roxolana Ridel", "@roxolana.ridel", "IG", 6.67, "日系动漫；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，个人写真式cos内容维持垂类关注。", HYPEAUDITOR, er=1.77),
    rec("Cosplay Jakarta", "@cosplayjakarta", "IG", 6.47, "漫展现实感；日系动漫", "广告合作", "因为互动人设+IP红利，地区cosplay社群通过漫展现场与角色内容聚合粉丝。", HYPEAUDITOR, er=4.2),
    rec("Komori Cosplay Costume", "@komoricosplaycostume", "IG", 6.27, "日系动漫；色气/性感风", "自营店铺；广告合作", "因为IP红利+强力视觉标签，服装账号用角色成衣展示承接coser购买需求。", HYPEAUDITOR, er=7.7),
    rec("puddingu", "@puddinguchan", "IG", 5.90, "日系动漫；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，小众二次元审美和高互动图文形成增长点。", HYPEAUDITOR, er=9.88),
    rec("David Ngo", "@dtjaaaam", "IG", 5.75, "漫展现实感；欧美游戏/影视", "广告合作", "因为互动人设+技术壁垒，摄影/漫展向cos内容连接创作者与粉丝社群。", HYPEAUDITOR, er=1.37),
    rec("Katy Andrea Cosplay", "@katy_garcia_39", "IG", 5.74, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为IP红利+互动人设，Miku等二次元角色与streamer身份叠加带来高互动。", HYPEAUDITOR, er=10.27, works="Miku（来源标题提及）；[缺失：其余最高互动角色]"),
    rec("Bane Armstrong", "@bane_armstrong", "IG", 9.02, "欧美游戏/影视；道具/机械装甲", "广告合作", "因为技术壁垒+IP红利，男性向游戏/影视角色与体格化表达形成差异化。", HYPEAUDITOR, er=4.35),
    rec("pro_cosplay", "@pro_cosplay", "IG", 16.03, "日系动漫；漫展现实感", "广告合作", "因为IP红利+互动人设，专业cosplay账号通过持续角色内容获得稳定垂类关注。", HYPEAUDITOR, er=4.09),
    rec("King Chris", "@imkingchristian", "IG", 370.0, "日系动漫；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，Dripkage人设把动漫梗做成可传播身份标签。", FEEDSPOT, notes="Feedspot公开粉丝数；近30日ER/涨粉未披露。"),
    rec("Jessica Nigri", "@jessicanigri", "IG", 360.0, "欧美游戏/影视；道具/机械装甲", "广告合作；自营店铺", "因为技术壁垒+IP红利+互动人设，复杂道具服装、游戏角色和多年线下声量共同放大影响力。", FEEDSPOT + "；" + HIREINFLUENCE, works="World of Warcraft；Monster Hunter（HireInfluence列举）"),
    rec("Ravengriim", "@ravengriim", "IG", 280.0, "洛丽塔/哥特/暗黑美学；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，恐怖与cosplay混合内容形成鲜明暗黑标签。", INFLDB),
    rec("HEIDI LAVON", "@heidilavon", "双平台", 250.0, "原创/OC；色气/性感风", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+强力视觉标签，YouTuber/Twitch Partner身份让cosplay内容获得跨平台扩散。", INFLDB),
    rec("Lonelyman", "@lowcostcosplayth", "IG", 200.0, "原创/OC；漫展现实感", "广告合作", "因为高频整活/短变装节奏+强力原创人设，低成本cosplay反差梗具备强分享属性。", INFLDB),
    rec("kameaam", "@kameaam", "IG", 180.0, "日系动漫；漫展现实感", "广告合作", "因为IP红利+互动人设，印尼动漫/游戏cosplay内容与本地粉丝社群连接紧密。", INFLDB),
    rec("Brandon Rogers", "@brandonbored", "IG", 170.0, "原创/OC；欧美游戏/影视", "广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，角色表演型内容把cosplay扩大到喜剧叙事。", INFLDB),
    rec("lyra crow", "@lyracr0w0", "IG", 170.0, "原创/OC；色气/性感风", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，猫梗/二次元审美和个人人设提升粉丝黏性。", INFLDB),
    rec("Alodia Quimbo", "@alodia", "双平台", 160.0, "日系动漫；欧美游戏/影视", "广告合作；直播收入（Twitch/抖音）", "因为IP红利+互动人设，游戏、动漫cosplay和东南亚电竞/娱乐资源叠加形成长期影响力。", FEEDSPOT + "；" + HIREINFLUENCE),
    rec("Dessyyc", "@oxdessyxo", "IG", 150.0, "日系动漫；色气/性感风", "广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，短视频式角色包装适合平台推荐。", FEEDSPOT, notes="Feedspot披露2mo +3%，未转换为30日涨粉。"),
    rec("Alina Becker", "@japp_leack", "IG", 150.0, "日系动漫；欧美游戏/影视", "广告合作", "因为IP红利+强力视觉标签，高质量动漫/游戏角色写真稳定承接热门IP关注。", FEEDSPOT + "；" + HIREINFLUENCE, works="动漫/游戏角色（HireInfluence泛称，未列最高互动单条）"),
    rec("Myrtle Sarrosa", "@myrtlegail", "双平台", 150.0, "日系动漫；欧美游戏/影视", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，演员、游戏主播和cosplay身份叠加，扩大菲律宾及东南亚影响力。", FEEDSPOT),
    rec("snitchery", "@snitchery", "IG", 150.0, "仿妆/特效妆；原创/OC", "广告合作", "因为强力原创人设/视觉美学标签+技术壁垒，妆造和怪诞审美让角色变身内容更易出圈。", INFLDB),
    rec("Taryn", "@taryn_cosplay", "IG", 140.0, "欧美游戏/影视；道具/机械装甲", "广告合作；自营店铺", "因为技术壁垒+互动人设，职业cosplayer和创业者身份增强漫展/品牌合作可信度。", FEEDSPOT),
    rec("Cinnanoe", "@cinnannoe", "IG", 140.0, "日系动漫；色气/性感风", "广告合作", "因为强力视觉标签+IP红利，anime/游戏审美和角色化写真构成稳定吸粉入口。", FEEDSPOT, notes="Feedspot披露2mo -2%，未转换为30日涨粉。"),
    rec("Rolyat", "@rolyat", "IG", 130.0, "欧美游戏/影视；日系动漫", "广告合作；直播收入（Twitch/抖音）", "因为IP红利+互动人设，cosplayer/streamer身份让游戏角色内容与直播社群互相导流。", FEEDSPOT),
    rec("Liz Katz", "@lizkatzofficial", "IG", 130.0, "欧美游戏/影视；色气/性感风", "广告合作", "因为IP红利+强力视觉标签，超级英雄、游戏与健身元素结合，形成娱乐化角色形象。", FEEDSPOT + "；" + HIREINFLUENCE, works="超级英雄角色（HireInfluence泛称）"),
    rec("Senyamiku", "@senyamiku", "双平台", 130.0, "日系动漫；色气/性感风", "广告合作", "因为IP红利+高频整活/短变装节奏，多平台账号矩阵放大中国二次元cosplay内容。", FEEDSPOT, notes="Feedspot披露2mo -1%，未转换为30日涨粉。"),
    rec("Kamui Cosplay", "@kamuicosplay", "IG", 130.0, "道具/机械装甲；欧美游戏/影视", "自营店铺；广告合作", "因为技术壁垒+IP红利，盔甲道具教程、书籍和游戏角色复刻构成高门槛护城河。", INFLDB + "；" + HIREINFLUENCE, works="Monster Hunter；The Witcher（HireInfluence列举）"),
    rec("moe_five", "@moe_five", "IG", 130.0, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；自营店铺", "因为强力视觉标签+互动人设，cosplayer/streamer身份结合Fanicon与通販形成粉丝转化。", INFLDB),
    rec("Alyson Tabbitha", "@alysontabbitha", "双平台", 120.0, "仿妆/特效妆；欧美游戏/影视", "广告合作；自营店铺", "因为技术壁垒+IP红利，仿妆、手作和影视角色变身教程拥有极强可验证辨识度。", FEEDSPOT + "；" + HIREINFLUENCE, works="Wonder Woman；Star Wars（HireInfluence列举）"),
    rec("BellaRama", "@bellaramatv", "IG", 100.0, "原创/OC；色气/性感风", "广告合作", "因为强力视觉标签+互动人设，洛杉矶内容创作者身份与角色化内容提升商业合作面。", INFLDB),
    rec("Valeriia Makusheva", "@leraferal", "IG", 100.0, "原创/OC；洛丽塔/哥特/暗黑美学", "广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，视觉风格清晰利于平台识别。", INFLDB),
    rec("Fukuro", "@f.ukuro", "IG", 99.31, "仿妆/特效妆；道具/机械装甲", "广告合作", "因为技术壁垒+IP红利，MUA与crafter身份让角色还原从妆面延伸到服装制作。", FEEDSPOT),
    rec("Takomayuyi", "@takomayuyi", "双平台", 93.12, "日系动漫；色气/性感风", "广告合作", "因为IP红利+高频整活/短变装节奏，中文/英文/日文多语种内容扩大跨地区传播。", FEEDSPOT),
    rec("Ely", "@eeelyeee", "IG", 91.86, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+IP红利，日系写真式cosplay与付费写真收藏订阅构成变现闭环。", FEEDSPOT),
    rec("Caitlin Christine", "@caitlinchristinee", "IG", 91.18, "欧美游戏/影视；仿妆/特效妆", "广告合作", "因为IP红利+技术壁垒，Master of disguise定位强化超级英雄/影视角色变身记忆点。", INFLDB),
    rec("Danny Phantom", "@danny.phantom.exe", "双平台", 89.95, "日系动漫；原创/OC", "自营店铺；广告合作", "因为强力原创人设/视觉美学标签+高频整活/短变装节奏，TikTok与merch导流支持角色化品牌。", INFLDB),
    rec("Jake Myers", "@polskihussar", "IG", 88.55, "欧美游戏/影视；道具/机械装甲", "广告合作", "因为IP红利+技术壁垒，Star Wars/Transformers玩具与服装展示连接硬核粉丝。", INFLDB),
    rec("say_catt", "@say_catt", "IG", 82.31, "日系动漫；色气/性感风", "广告合作", "因为强力视觉标签+IP红利，巴西cosplay账号通过动漫/游戏角色获得跨语种粉丝。", INFLDB),
    rec("Luxlo", "@luxlocosplay", "双平台", 77.94, "原创/OC；洛丽塔/哥特/暗黑美学", "直播收入（Twitch/抖音）；广告合作", "因为互动人设+强力视觉标签，Twitch streamer与fairy审美标签共同强化粉丝黏性。", FEEDSPOT),
    rec("Ethan / EJ", "@p4perback", "双平台", 73.28, "欧美游戏/影视；日系动漫", "广告合作", "因为高频整活/短变装节奏+IP红利，TikTok 2.2M体量与IG角色内容互相导流。", INFLDB),
    rec("Jo Steel", "@artistjodysteel", "IG", 71.44, "仿妆/特效妆；原创/OC", "广告合作", "因为技术壁垒+强力原创人设，makeup artist身份使cosplay、body art和角色妆形成差异化。", INFLDB),
    rec("Xiao Yukiko", "@xiaoyukiko", "IG", 70.25, "日系动漫；色气/性感风", "广告合作", "因为IP红利+强力视觉标签，新加坡cosplay内容通过PR/Event合作扩展商业曝光。", FEEDSPOT),
    rec("Alice Delish", "@alicedelish", "IG", 68.50, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+IP红利，cosplay与makeup结合并通过link聚合导流粉丝。", FEEDSPOT),
    rec("chihiro_chang", "@chihiro_chang", "IG", 67.25, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，台湾cosplayer/VTuber/女团身份叠加，扩展多场景粉丝触点。", INFLDB),
    rec("Anni The Duck", "@annitheduck", "双平台", 65.08, "日系动漫；欧美游戏/影视", "直播收入（Twitch/抖音）；广告合作", "因为互动人设+高频整活/短变装节奏，streamer persona与cosplay内容形成强社群互动。", INFLDB),
    rec("LittleJem", "@littlejem", "双平台", 65.29, "道具/机械装甲；欧美游戏/影视", "广告合作；自营店铺", "因为技术壁垒+互动人设，costume/props maker身份与YouTube大体量教程内容互相强化。", FEEDSPOT),
    rec("seracoss", "@seracoss", "IG", 63.32, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+互动人设，kofi会员请求机制让角色内容和粉丝付费需求相连。", INFLDB),
    rec("punipun7", "@punipun7", "IG", 62.83, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，cosplay、game、music与品牌大使身份共同带动印尼圈层传播。", INFLDB),
    rec("Danielle Denicola", "@danielledenicola", "双平台", 62.0, "欧美游戏/影视；日系动漫", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，cosplayer/Twitch streamer/voice actor身份覆盖游戏、播客和角色内容。", FEEDSPOT),
    rec("Maul", "@maul_cosplay", "IG", 61.97, "欧美游戏/影视；道具/机械装甲", "广告合作", "因为技术壁垒+IP红利，stuntman和crafting全年输出强化游戏/影视角色真实感。", INFLDB),
    rec("Lani", "@itslanicos", "双平台", 60.88, "日系动漫；色气/性感风", "广告合作", "因为互动人设+高频整活/短变装节奏，TikTok导流和漫展行程让粉丝持续追踪。", FEEDSPOT),
    rec("K A T I E", "@katie_westwood", "IG", 60.53, "日系动漫；原创/OC", "广告合作", "因为强力视觉标签+IP红利，日本所在地与日系角色内容强化二次元语境。", INFLDB),
    rec("Meg Turney", "@megturney", "双平台", 60.13, "日系动漫；欧美游戏/影视", "直播收入（Twitch/抖音）；广告合作", "因为互动人设+IP红利，Twitch streamer、cosplayer与Gundam/游戏内容形成老粉基础。", FEEDSPOT),
    rec("Yaya Han", "@yayahan", "IG", 59.16, "道具/机械装甲；欧美游戏/影视", "自营店铺；广告合作", "因为技术壁垒+互动人设，设计师/作者身份把cosplay从作品扩展到面料、教程和行业标准。", FEEDSPOT + "；" + HIREINFLUENCE),
    rec("PanickedCosplays", "@panickedcosplays", "TikTok", 59.10, "日系动漫；漫展现实感", "广告合作", "因为高频整活/短变装节奏+互动人设，TikTok动画/cosplay类目排名与持续互动带来推荐流曝光。", SEARCH, notes="StarNgage公开片段：591K TikTok followers；未披露ER。"),
    rec("Kora Aura", "@koraaura", "双平台", 57.45, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，cosplayer/streamer身份和漫展行程维系高黏性粉丝。", INFLDB),
    rec("Kinpatsu Cosplay", "@kinpatsucosplay", "IG", 55.44, "道具/机械装甲；欧美游戏/影视", "自营店铺；广告合作", "因为技术壁垒+IP红利，教程、纸样和高效率服装制作构成教育型影响力。", FEEDSPOT + "；" + HIREINFLUENCE),
    rec("styleitwithrj", "@styleitwithrj", "双平台", 54.28, "欧美游戏/影视；漫展现实感", "广告合作", "因为IP红利+高频整活/短变装节奏，Disney/Star Wars/Marvel等泛娱乐IP适合短视频扩散。", INFLDB),
    rec("thousandhunny", "@thousandhunny", "IG", 51.84, "日系动漫；色气/性感风", "广告合作", "因为强力视觉标签+IP红利，红发cosplayer形象和角色写真提供清晰记忆点。", INFLDB),
    rec("Nathan Weir", "@nate.k.weir", "双平台", 51.12, "欧美游戏/影视；原创/OC", "广告合作", "因为IP红利+高频整活/短变装节奏，fandoms/fitness组合与TikTok 2M+导流拓宽受众。", INFLDB),
    rec("Mimi Semaan", "@mimisemaan", "双平台", 50.49, "日系动漫；原创/OC", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+强力视觉标签，cosplayer/streamer/fitness身份让角色内容拥有生活方式入口。", INFLDB),
    rec("Karrigan Taylor", "@thekarrigantaylor", "IG", 50.62, "洛丽塔/哥特/暗黑美学；原创/OC", "广告合作；自营店铺", "因为强力原创人设/视觉美学标签+互动人设，succubus/alt fashion和音乐身份形成暗黑美学品牌。", INFLDB),
    rec("akrcos", "@akrcos", "IG", 50.41, "日系动漫；原创/OC", "广告合作", "因为互动人设+IP红利，writer/cosplayer定位让角色内容兼具叙事与粉丝沟通。", INFLDB),
    rec("Valentina Kryp", "@vkryp", "双平台", 49.32, "日系动漫；欧美游戏/影视", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，cosplayer/streamer/host身份覆盖拉美游戏和漫展场景。", INFLDB),
    rec("Halcybella", "@halcybella", "IG", 48.55, "原创/OC；日系动漫", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，shapeshifter定位强化多角色切换期待。", INFLDB),
    rec("Yuji Koi", "@yuji.hannah", "IG", 47.58, "道具/机械装甲；欧美游戏/影视", "广告合作", "因为技术壁垒+互动人设，WCS/ECG奖项与3D打印品牌大使身份提升专业信任。", INFLDB),
    rec("Raine Emery", "@raineemery", "IG", 47.03, "道具/机械装甲；原创/OC", "广告合作", "因为技术壁垒+强力原创人设，costume designer和fashion design背景提高服装完成度。", INFLDB),
    rec("Adina Luna", "@adinascozylife", "IG", 46.31, "原创/OC；日系动漫", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，cozy fashion与cosplay混合内容扩大生活方式受众。", INFLDB),
    rec("leonchiro", "@leonchiro", "IG", 45.35, "欧美游戏/影视；道具/机械装甲", "广告合作", "因为技术壁垒+IP红利，职业shapeshifter与摔角/体能训练增强游戏角色真实感。", INFLDB),
    rec("Michael Slayers", "@michaelslayersofficial", "IG", 45.11, "欧美游戏/影视；原创/OC", "广告合作", "因为IP红利+强力原创人设，角色化个人账号依靠影视/游戏审美维持关注。", INFLDB),
    rec("edelynlyn", "@edelynlyn", "IG", 43.95, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，印尼游戏、健身和cosplay资源连接本地粉丝。", INFLDB),
    rec("Milkiverse", "@milkiwifey", "IG", 43.04, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+互动人设，waifu/gamer/nerd定位与私域内容导流形成转化路径。", INFLDB),
    rec("Hendo", "@hendoart", "IG", 42.37, "道具/机械装甲；欧美游戏/影视", "广告合作", "因为技术壁垒+IP红利，props制作和角色摄影展示兼顾教程价值与视觉冲击。", INFLDB),
    rec("onisuna", "@onisuna", "双平台", 40.57, "日系动漫；原创/OC", "直播收入（Twitch/抖音）；广告合作", "因为互动人设+IP红利，Twitch partner与cosplayer身份让直播社群反哺IG内容。", INFLDB),
    rec("kurusu_usako", "@kurusu_usako", "双平台", 40.54, "日系动漫；色气/性感风", "自营店铺；广告合作", "因为强力视觉标签+互动人设，cosplayer/YouTuber与写真/DVD/服饰线索构成多渠道变现。", INFLDB),
    rec("Jennings Brower", "@jenningsbrower", "IG", 39.82, "欧美游戏/影视；漫展现实感", "广告合作", "因为IP红利+互动人设，multiverse定位和LA影像化内容适合超级英雄角色传播。", INFLDB),
    rec("Saya Fox", "@sayathefox", "IG", 38.96, "日系动漫；色气/性感风", "广告合作", "因为强力视觉标签+IP红利，马来西亚cosplayer通过不断换装强化角色期待。", INFLDB),
    rec("Ran", "@weilanran", "双平台", 38.66, "日系动漫；色气/性感风", "广告合作", "因为IP红利+高频整活/短变装节奏，多平台中文cosplayer矩阵扩大海外触达。", FEEDSPOT, notes="Feedspot披露2mo -2%，未转换为30日涨粉。"),
    rec("mizterfilip", "@mizterfilip", "双平台", 36.77, "日系动漫；原创/OC", "广告合作", "因为IP红利+互动人设，日语歌唱、插画和YouTube内容使cosplay账号具备多兴趣入口。", INFLDB),
    rec("kurumipurarine", "@kurumipurarine", "IG", 36.46, "日系动漫；色气/性感风", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，cosplayer/DJ/singer身份与海外出演经历扩大线下影响力。", INFLDB),
    rec("kurasaka_kururu", "@kurasaka_kururu", "IG", 35.63, "日系动漫；色气/性感风", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+互动人设，Fantia写真与线下会面日程支持粉丝付费转化。", INFLDB),
    rec("Millie", "@whatsername1.0", "IG", 35.42, "原创/OC；日系动漫", "广告合作", "因为互动人设+强力原创人设，苏格兰本地化表达和cosplay日常内容降低粉丝距离感。", INFLDB),
    rec("Gale F", "@allthepotsnpans.cos", "IG", 35.35, "日系动漫；漫展现实感", "广告合作", "因为互动人设+IP红利，NYC本地cosplayer通过漫展行程维系可见度。", INFLDB),
    rec("slightlyvillainous", "@slightlyvillainous", "双平台", 34.73, "欧美游戏/影视；原创/OC", "广告合作", "因为高频整活/短变装节奏+互动人设，TikTok 5M与YouTube 1M为IG持续导流。", INFLDB),
    rec("Miss Twisted", "@misstwisted", "双平台", 34.41, "原创/OC；日系动漫", "广告合作", "因为高频整活/短变装节奏+互动人设，YouTube与澳洲本地cosplay内容提供持续曝光。", INFLDB),
    rec("deadpudds", "@deadpudds", "双平台", 33.84, "日系动漫；漫展现实感", "广告合作", "因为互动人设+IP红利，唱歌、视频和cosplay内容结合，让角色内容更贴近日常娱乐。", INFLDB),
    rec("Mauro Rossiello", "@maurorossielloofficial", "IG", 32.61, "原创/OC；色气/性感风", "广告合作", "因为强力视觉标签+互动人设，鲜明个人形象和商业邮箱导向支持品牌合作。", INFLDB),
    rec("kiyocosplay", "@kiyocosplay", "双平台", 32.12, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，professional cosplayer/streamer/meme queen定位兼具角色与社交娱乐。", INFLDB),
    rec("ceekayecosplays", "@ceekayecosplays", "IG", 27.80, "道具/机械装甲；日系动漫", "广告合作", "因为技术壁垒+互动人设，professional cosplayer/cosmaker身份让作品制作过程具备可信度。", INFLDB),
    rec("Sakuramomo", "@skauramomo", "IG", 27.62, "洛丽塔/哥特/暗黑美学；日系动漫", "广告合作", "因为强力视觉标签+互动人设，全职coser与可爱梦幻少女洛丽塔风格形成清晰定位。", FEEDSPOT),
    rec("aliya.will", "@aliya.will", "IG", 26.90, "原创/OC；日系动漫", "广告合作", "因为强力原创人设/视觉美学标签+IP红利，artist/cosplay定位让角色内容兼具画面和设定感。", INFLDB),
    rec("Laila", "@mishirudo", "IG", 26.77, "日系动漫；色气/性感风", "广告合作", "因为IP红利+互动人设，日本事务所/活动信息让角色写真与线下曝光相互增强。", INFLDB),
    rec("Pandorya", "@pandorya", "IG", 26.31, "日系动漫；原创/OC", "自营店铺；广告合作", "因为互动人设+IP红利，manga推广和角色化社群内容支持粉丝长期关注。", INFLDB),
    rec("helene.tw", "@helene.tw", "IG", 25.82, "日系动漫；漫展现实感", "广告合作", "因为互动人设+IP红利，台湾cosplayer通过公开活动与动漫展摊位形成线下触点。", INFLDB),
    rec("Torra", "@dystorra", "IG", 25.79, "日系动漫；原创/OC", "付费订阅（OF/Patreon）；广告合作", "因为强力视觉标签+互动人设，character select定位与私域内容链接形成角色期待。", INFLDB),
    rec("Darshelle Stevens", "@darshellestevens", "IG", 25.60, "欧美游戏/影视；原创/OC", "广告合作", "因为强力视觉标签+IP红利，摄影级角色内容和长期cosplay形象保持商业辨识度。", FEEDSPOT),
    rec("abab_abab_0", "@abab_abab_0", "双平台", 25.07, "日系动漫；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为高频整活/短变装节奏+IP红利，TikTok 71万与cosplay内容互相导流。", INFLDB),
    rec("derailed3d", "@derailed3d", "双平台", 24.79, "道具/机械装甲；欧美游戏/影视", "付费订阅（OF/Patreon）；广告合作", "因为技术壁垒+高频整活/短变装节奏，3D工程艺术项目具备强过程感和平台传播性。", INFLDB),
    rec("Linzor", "@linzor", "IG", 24.56, "日系动漫；原创/OC", "广告合作；自营店铺", "因为互动人设+强力原创人设，Scandyswe创始人身份强化北欧cosplay社群影响力。", INFLDB),
    rec("hamu_cotton", "@hamu_cotton", "IG", 23.92, "日系动漫；色气/性感风", "广告合作；自营店铺", "因为强力视觉标签+互动人设，日英双语商务和活动信息支持跨地区粉丝转化。", INFLDB),
    rec("Madison Nicole", "@zoogirlq", "IG", 22.90, "原创/OC；日系动漫", "广告合作", "因为强力原创人设/视觉美学标签+互动人设，STEM/角色扮演反差设定增强个人记忆点。", INFLDB),
    rec("valiantf0x", "@valiantf0x", "IG", 22.86, "欧美游戏/影视；漫展现实感", "广告合作", "因为互动人设+IP红利，经纪代表与角色扮演内容提升商务与线下活动可见度。", INFLDB),
    rec("sunflow3rsamurai", "@sunflow3rsamurai", "双平台", 22.55, "日系动漫；原创/OC", "广告合作", "因为IP红利+互动人设，artist/cosplayer/BGM身份和TikTok 300K为IG提供导流。", INFLDB),
    rec("infamousbylaura", "@infamousbylaura", "双平台", 22.19, "欧美游戏/影视；漫展现实感", "广告合作；直播收入（Twitch/抖音）", "因为互动人设+IP红利，悉尼cosplay、Twitch和YouTube内容维系多平台粉丝关系。", INFLDB),
    rec("xier_cos", "@xier_cos", "双平台", 22.08, "日系动漫；漫展现实感", "广告合作", "因为IP红利+互动人设，波士顿华语cosplayer通过TT/X同名矩阵扩大触达。", INFLDB),
    rec("Kade", "@dinograveyard", "双平台", 21.90, "原创/OC；欧美游戏/影视", "广告合作", "因为强力原创人设+高频整活/短变装节奏，cosplay/drag/storytelling与TikTok/YouTube百万级导流结合。", INFLDB),
]


def er_value(row):
    return row["er"] if isinstance(row["er"], (int, float)) else None


def score(row):
    er = er_value(row)
    return row["followers"] * er if er is not None else None


def display_number(value):
    if value is None:
        return MISSING
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


ranked = sorted(
    records,
    key=lambda row: (
        score(row) is not None,
        score(row) if score(row) is not None else -1,
        row["followers"],
    ),
    reverse=True,
)[:100]


headers = [
    "排名",
    "博主名称",
    "主要平台",
    "粉丝数(万)",
    "互动率(%)",
    "涨粉率(%)",
    "主要风格",
    "变现方式",
    "爆款代表作",
    "火的原因(一句话)",
]


def row_values(index, row):
    return [
        index,
        f"{row['name']} ({row['handle']})",
        row["platform"],
        display_number(row["followers"]),
        display_number(row["er"]),
        display_number(row["growth"]),
        row["style"],
        row["monetization"],
        row["works"],
        row["reason"],
    ]


md_lines = [
    "# 全球 Cosplay 博主 Top 100 分析报告",
    "",
    f"- 采集日期：{REPORT_DATE}",
    "- 排序口径：优先按可验证 `粉丝数(万) × 公开互动率(%)` 排序；互动率缺失账号不填估算值，置于可计算账号之后并按粉丝数降序。",
    "- 缺失规则：公开页面、第三方列表与搜索片段均无法验证近30日字段时，按要求标注 `[缺失；尝试3次]`。",
    "- 主要来源：HypeAuditor、Feedspot、InflDB、HireInfluence、Collabstr、公开搜索结果片段。",
    "",
    "## 报告 A：Markdown 表格",
    "",
    "|" + "|".join(headers) + "|",
    "|" + "|".join(["---"] * len(headers)) + "|",
]

for idx, row in enumerate(ranked, 1):
    values = [str(v).replace("\n", " ") for v in row_values(idx, row)]
    md_lines.append("|" + "|".join(values) + "|")

deep_by_handle = {
    "@knitecoser": "KNITE 的爆火路径更接近“专业造型+跨平台角色品牌”。公开第三方片段显示其粉丝量和互动率都处于高位，说明不是单纯靠粉丝基数，而是靠游戏/动漫角色的高完成度与持续互动把受众沉淀下来。其优势在于技术壁垒明显，角色视觉具备强辨识度，同时可以承接广告、活动和自营内容，形成从作品到商业合作的闭环。",
    "@cosplayshorty00": "cosplayshorty00 更偏 TikTok 友好的增长模型：短视频高频、角色切换快、视觉刺激强。公开片段显示其视频互动率较高，这类账号的关键不在长篇叙事，而在几秒内完成变装、表情和角色钩子。日系动漫与性感风标签让内容容易被推荐系统识别，也更容易被同好二创、转发和收藏。",
    "@komori198": "Komori Cosplay 的优势是稳定二次元视觉和商业化路径清晰。HypeAuditor 公开页给出较高互动率，说明其粉丝不是纯沉默关注。账号同时关联 cosplay costume 合作，代表其内容不仅是个人写真，也能承接服装展示和购买转化。她的走红密码是热门动漫 IP、统一审美与商品化链路叠加。",
    "@pinku.cosplay": "PINKU cosplay 在中腰部账号里互动率突出，走红点是非常清晰的粉色系/二次元视觉标签。相比只做热门角色复刻的账号，她更像把角色内容包装成个人审美品牌：粉丝记住的不只是某个 IP，而是账号整体氛围。IP 红利带来入口，强视觉标签负责沉淀关注和复访。",
    "@fabibiworldcosplay": "Fabibi World Cosplay 的差异化来自 cosplay 与旅行、生活化场景的结合。她不是只在棚拍里完成角色，而是把角色形象放进更轻松的内容语境中，降低非核心粉丝的观看门槛。公开数据源显示其粉丝规模和互动率都有可观基础，说明泛娱乐受众与cosplay垂类粉丝可以同时被覆盖。",
    "@bellascosplay": "bellascosplay 的增长逻辑是垂类聚焦：账号名、内容方向和受众预期都非常明确。HypeAuditor 显示其粉丝规模高于多数同类账号且互动率不低，说明“动漫+cosplay girl”的直接定位仍然有效。她的关键不是复杂叙事，而是持续输出二次元角色与统一视觉风格，长期占据粉丝心智。",
    "@rhauura": "Nabi 的爆火路径偏暗黑二次元审美。她把洛丽塔、哥特和动漫角色结合起来，形成比普通角色复刻更鲜明的视觉差异。HypeAuditor 公开互动率显示该审美能带来较高粉丝回应。此类账号的优势是可被一眼识别，适合在图文、Reels和粉丝收藏场景中沉淀长期关注。",
    "@uwowo.cosplay": "Uwowo Cosplay Official 是品牌型 cosplay 账号，走红密码不是单一博主人设，而是“热门 IP 服装展示+购买需求”。HypeAuditor 显示其互动率具备竞争力，说明商品账号也能通过角色内容获得社媒传播。其核心优势在于把粉丝兴趣直接连接到自营店铺，内容即橱窗，角色即商品入口。",
    "@ribaibu": "Ribaibu Cosplay 的优势在于个人 coser 身份清晰，内容既有角色还原也有社群互动。她的粉丝量不算最大，但公开互动率处于较好水平，说明受众黏性强。走红路径更依赖持续角色输出、漫展/粉丝互动和可亲近的人设，而不是一次性爆款。对品牌而言，这类账号适合做垂类深度触达。",
    "@katy_garcia_39": "Katy Andrea Cosplay 的公开互动率很高，且标题直接关联 Miku cosplay 与 streamer 身份。她的爆火机制是初音未来等强 IP 带来搜索和推荐入口，再由直播/互动人设提升粉丝关系密度。小体量但高互动说明她的内容更偏高黏性圈层，适合动漫、直播和角色周边相关合作。",
}

md_lines += ["", "## 报告 B：前 10 位的深度分析", ""]
for idx, row in enumerate(ranked[:10], 1):
    md_lines.append(f"### {idx}. {row['name']} ({row['handle']})")
    md_lines.append(deep_by_handle.get(row["handle"], row["reason"]))
    md_lines.append("")

md_lines += [
    "## 数据来源与缺失说明",
    "",
    "- HypeAuditor 公开页提供部分账号 follower 与 engagement rate，但未公开近30日涨粉率。",
    "- Feedspot/InflDB 公开页提供大量账号粉丝数、简介、部分2个月涨粉字段；2个月涨粉未换算为近30日涨粉。",
    "- IG/TikTok 官方公开接口在无登录/API授权场景下未返回每账号近30日 media insights，因此 ER/涨粉率不做估算。",
    "- 最高播放/点赞作品只有在公开来源明确列出时填写；否则标注缺失。",
]

md_path = OUT_DIR / "global_cosplay_top100_report.md"
md_path.write_text("\n".join(md_lines), encoding="utf-8")

wb = Workbook()
ws = wb.active
ws.title = "报告A_Top100"
ws.append(headers)
for idx, row in enumerate(ranked, 1):
    ws.append(row_values(idx, row))

source_ws = wb.create_sheet("数据来源_缺失说明")
source_ws.append([
    "排名",
    "账号",
    "来源URL或依据",
    "粉丝数来源",
    "ER/涨粉/近30日更新说明",
    "加权分",
])
for idx, row in enumerate(ranked, 1):
    source_ws.append([
        idx,
        f"{row['name']} ({row['handle']})",
        row["source"],
        f"{row['followers']}万；来源见URL/依据",
        row["notes"] or "近30日ER、近30日涨粉率、官方insights未公开；尝试：第三方页、官方公开页、搜索片段。",
        score(row) if score(row) is not None else "[缺失]",
    ])

deep_ws = wb.create_sheet("报告B_前10深度分析")
deep_ws.append(["排名", "账号", "深度分析"])
for idx, row in enumerate(ranked[:10], 1):
    deep_ws.append([idx, f"{row['name']} ({row['handle']})", deep_by_handle.get(row["handle"], row["reason"])])

meta_ws = wb.create_sheet("采集口径")
meta = [
    ["采集日期", REPORT_DATE],
    ["生成时间UTC", datetime.now(UTC).isoformat(timespec="seconds")],
    ["排序口径", "可验证ER账号按粉丝数(万)*ER(%)；ER缺失账号置后按粉丝数降序。"],
    ["缺失标注", MISSING],
    ["公开来源", "\n".join([HYPEAUDITOR, FEEDSPOT, INFLDB, HIREINFLUENCE, COLLABSTR])],
    ["官方接口限制", "未使用登录态、私有接口或需要授权的insights；无法公开验证近30日media insights。"],
]
for row in meta:
    meta_ws.append(row)

for sheet in wb.worksheets:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value[:80]) + 2), 60)
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

xlsx_name = "global_cosplay_top100_report.xlsx"
xlsx_path = OUT_DIR / xlsx_name
desktop_path = DESKTOP / xlsx_name
wb.save(xlsx_path)
wb.save(desktop_path)

print(f"Markdown: {md_path.resolve()}")
print(f"Excel: {xlsx_path.resolve()}")
print(f"Desktop Excel: {desktop_path.resolve()}")
